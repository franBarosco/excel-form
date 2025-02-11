import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
import re
import os

from openpyxl.reader.excel import load_workbook

file_name = 'data.xlsx'
#Check if the file already exist
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
else:
    #Make excel sheet
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Age', 'Mail', 'Phone', 'Address'])

def save_data():
    name = entry_name.get()
    age = entry_age.get()
    mail = entry_mail.get()
    phone = entry_phone.get()
    address = entry_address.get()

    #Check all the fields
    if not name or not age or not mail or not phone or not address:
        messagebox.showwarning('Warning', 'All fields are required')
        return

    #Check phone and age
    try:
        age = int(age)
        phone = int(phone)
    except ValueError:
        messagebox.showwarning('Warning', 'Age and Phone fields must be numbers')

    #Check e-mail
    if not re.match(r'[^@]+@[^@]+\.[^@]', mail):
        messagebox.showwarning('Warning', 'Please, enter a valid e-mail')
        return

    ws.append([name, age, mail, phone, address])
    wb.save(file_name)
    messagebox.showinfo('Perfect', 'Data saved')

    #Clear the data
    entry_name.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    entry_mail.delete(0, tk.END)
    entry_phone.delete(0, tk.END)
    entry_address.delete(0, tk.END)


#Initialize tkinter
root = tk.Tk()
root.title("Data Form")
root.configure(bg='#3D8D7A')
label_style = {"bg": '#3D8D7A', "fg": 'white'}
entry_style = {"bg": '#FBFFE4', "fg": 'Black'}

#Creating labels

#Name
label_name = tk.Label(root, text='Name', **label_style)
label_name.grid(row=0, column=0, padx=10, pady=5)
entry_name = tk.Entry(root, **entry_style)
entry_name.grid(row=0, column=1, padx=10, pady=5)

#Age
label_age = tk.Label(root, text='Age', **label_style)
label_age.grid(row=1, column=0, padx=10, pady=5)
entry_age = tk.Entry(root, **entry_style)
entry_age.grid(row=1, column=1, padx=10, pady=5)

#mail
label_mail = tk.Label(root, text='Mail', **label_style)
label_mail.grid(row=2, column=0, padx=10, pady=5)
entry_mail = tk.Entry(root, **entry_style)
entry_mail.grid(row=2, column=1, padx=10, pady=5)

#phone
label_phone = tk.Label(root, text='Phone', **label_style)
label_phone.grid(row=3, column=0, padx=10, pady=5)
entry_phone = tk.Entry(root, **entry_style)
entry_phone.grid(row=3, column=1, padx=10, pady=5)

#address
label_address = tk.Label(root, text='Address', **label_style)
label_address.grid(row=4, column=0, padx=10, pady=5)
entry_address = tk.Entry(root, **entry_style)
entry_address.grid(row=4, column=1, padx=10, pady=5)

#Save Button
save_button = tk.Button(root, text="Save", command=save_data, bg='#B3D8A8', fg="black", width=10)
save_button.grid(row=5, column=0, columnspan=2, padx=10, pady=10)


root.mainloop()
